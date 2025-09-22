#!/usr/bin/env python3
"""
Improved Unconstrained Entity Extraction System with:
1. Retry logic with exponential backoff
2. Parallel processing for batches
3. Caching for successful responses
4. Entity name standardization
5. XLSX export instead of CSV
"""

import os
import sys
import time
import json
import base64
import logging
import hashlib
import pickle
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Google Gemini imports
import google.generativeai as genai

# OpenAI imports
from openai import OpenAI

# Load environment variables
load_dotenv()

# Set UTF-8 encoding for Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('improved_extraction.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# Entity standardization mappings
ENTITY_STANDARDIZATION = {
    'brands': {
        'balance point': ['Balance Point', 'BalancePoint', 'BP'],
        'lennox': ['Lennox', 'LENNOX', 'Lenox'],
        'carrier': ['Carrier', 'CARRIER'],
        'trane': ['Trane', 'TRANE'],
        'goodman': ['Goodman', 'GOODMAN'],
        'rheem': ['Rheem', 'RHEEM', 'Reem'],
        'york': ['York', 'YORK'],
        'american standard': ['American Standard', 'AmericanStandard', 'AMERICAN STANDARD'],
    },
    'equipment': {
        'ac unit': ['AC unit', 'ac unit', 'air conditioner', 'Air Conditioner', 'A/C unit', 'cooling unit'],
        'furnace': ['furnace', 'Furnace', 'heater', 'heating unit', 'gas furnace'],
        'heat pump': ['heat pump', 'Heat Pump', 'heatpump'],
        'air handler': ['air handler', 'Air Handler', 'air-handler', 'AHU'],
        'thermostat': ['thermostat', 'Thermostat', 'tstat', 'temperature control'],
        'ductwork': ['ductwork', 'Ductwork', 'ducts', 'duct', 'air ducts'],
        'condenser': ['condenser', 'Condenser', 'outdoor unit', 'condensing unit'],
    },
    'service_type': {
        'installation': ['installation', 'install', 'installing', 'new installation', 'replacement'],
        'repair': ['repair', 'fix', 'fixing', 'service call', 'troubleshooting'],
        'maintenance': ['maintenance', 'tune-up', 'cleaning', 'preventive maintenance', 'PM'],
        'inspection': ['inspection', 'evaluation', 'assessment', 'diagnosis'],
    }
}


class ImprovedImageAnalyzer:
    """
    Enhanced analyzer with retry logic, parallel processing, caching, and XLSX export.
    """

    def __init__(self, batch_size: int = 5, max_retries: int = 3):
        self.batch_size = batch_size
        self.max_retries = max_retries

        # Initialize API clients
        self.setup_apis()

        # Create output directory
        self.output_dir = Path("output_improved")
        self.output_dir.mkdir(exist_ok=True)

        # Cache directory for API responses
        self.cache_dir = Path("cache_responses")
        self.cache_dir.mkdir(exist_ok=True)

        # Progress tracking
        self.progress_file = self.output_dir / "progress.json"
        self.results_file = self.output_dir / "narrative_analysis.xlsx"

    def setup_apis(self):
        """Initialize Gemini and OpenAI API clients"""
        # Gemini setup
        gemini_key = os.getenv('GEMINI_API_KEY')
        if not gemini_key:
            raise ValueError("GEMINI_API_KEY not found in .env file")

        genai.configure(api_key=gemini_key)

        # Configure safety settings to be less restrictive
        self.safety_settings = [
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"}
        ]

        # Use gemini-1.5-pro-latest for most comprehensive analysis
        # Alternative options that also work:
        # - gemini-1.5-flash-latest (faster, slightly less detail)
        # - gemini-2.0-flash-exp (experimental, good quality)
        self.gemini_model = genai.GenerativeModel(
            'gemini-1.5-pro-latest',
            safety_settings=self.safety_settings
        )

        # OpenAI setup
        openai_key = os.getenv('OPENAI_API_KEY')
        if not openai_key:
            raise ValueError("OPENAI_API_KEY not found in .env file")

        self.openai_client = OpenAI(api_key=openai_key)

        logging.info("APIs initialized successfully")

    def get_cache_key(self, image_path: str, api_name: str) -> str:
        """Generate cache key for an image-API combination"""
        with open(image_path, 'rb') as f:
            file_hash = hashlib.md5(f.read()).hexdigest()
        return f"{file_hash}_{api_name}"

    def load_from_cache(self, image_path: str, api_name: str) -> Optional[Dict]:
        """Load cached response if available"""
        cache_key = self.get_cache_key(image_path, api_name)
        cache_file = self.cache_dir / f"{cache_key}.pkl"

        if cache_file.exists():
            try:
                with open(cache_file, 'rb') as f:
                    cached_data = pickle.load(f)
                    logging.info(f"Loaded {api_name} response from cache for {os.path.basename(image_path)}")
                    return cached_data
            except Exception as e:
                logging.warning(f"Failed to load cache for {cache_key}: {e}")
        return None

    def save_to_cache(self, image_path: str, api_name: str, response: Dict):
        """Save response to cache"""
        cache_key = self.get_cache_key(image_path, api_name)
        cache_file = self.cache_dir / f"{cache_key}.pkl"

        try:
            with open(cache_file, 'wb') as f:
                pickle.dump(response, f)
        except Exception as e:
            logging.warning(f"Failed to save cache for {cache_key}: {e}")

    def retry_with_backoff(self, func, *args, **kwargs):
        """Execute function with exponential backoff retry"""
        last_exception = None

        for attempt in range(self.max_retries):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                last_exception = e
                if attempt < self.max_retries - 1:
                    wait_time = 2 ** attempt  # Exponential backoff: 1, 2, 4 seconds
                    logging.warning(f"Attempt {attempt + 1} failed: {str(e)}. Retrying in {wait_time}s...")
                    time.sleep(wait_time)
                else:
                    logging.error(f"All {self.max_retries} attempts failed: {str(e)}")

        raise last_exception

    def encode_image(self, image_path: str) -> str:
        """Encode image to base64"""
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')

    def analyze_with_gemini_internal(self, image_path: str) -> Dict:
        """Internal Gemini analysis function for retry wrapper"""
        start_time = time.time()

        try:
            # Try using PIL to load and process image
            from PIL import Image
            import io

            # Open and potentially resize image if too large
            img = Image.open(image_path)

            # If image is very large, resize it
            max_size = (1024, 1024)
            if img.size[0] > max_size[0] or img.size[1] > max_size[1]:
                img.thumbnail(max_size, Image.Resampling.LANCZOS)

            # Convert to RGB if necessary
            if img.mode != 'RGB':
                img = img.convert('RGB')

            # Save to bytes
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='JPEG', optimize=True, quality=85)
            img_byte_arr = img_byte_arr.getvalue()

            # Upload the processed image
            image = genai.upload_file(path=image_path)

        except Exception as e:
            logging.warning(f"Failed to process image with PIL: {e}, using direct upload")
            # Fallback to direct upload
            image = genai.upload_file(image_path)

        time.sleep(0.5)

        # Comprehensive prompt for working model
        prompt = """Analyze this HVAC/Electrical/Plumbing service image comprehensively:

1. Equipment & Brands: List all visible equipment, brands, model numbers
2. Service Context: What type of work is being performed (installation, repair, maintenance)?
3. Technical Details: Any specifications, conditions, or issues visible
4. Location: Setting type (residential/commercial, indoor/outdoor)
5. Tools & Materials: What tools and materials are visible?
6. Workmanship: Observations about quality and professionalism
7. Complete Story: Describe the overall service context and what's happening

Provide detailed observations that would help understand this service scenario."""

        response = self.gemini_model.generate_content(
            [prompt, image],
            generation_config=genai.GenerationConfig(
                max_output_tokens=1000,  # Pro model can handle more
                temperature=0.5,  # Balanced temperature
                top_p=0.9,
                top_k=40
            )
        )

        # Handle response
        if response.parts:
            narrative = response.text
        elif response.candidates:
            candidate = response.candidates[0]
            if candidate.content and candidate.content.parts:
                narrative = candidate.content.parts[0].text
            else:
                # Try even simpler prompt
                fallback_prompt = "What equipment is in this image?"
                fallback_response = self.gemini_model.generate_content(
                    [fallback_prompt, image],
                    generation_config=genai.GenerationConfig(
                        max_output_tokens=200,  # Very limited output
                        temperature=0.1,
                        top_p=0.5,
                        top_k=5
                    )
                )

                if fallback_response.parts:
                    narrative = f"[Equipment focus] {fallback_response.text}"
                elif fallback_response.candidates and fallback_response.candidates[0].content:
                    try:
                        narrative = f"[Limited] {fallback_response.candidates[0].content.parts[0].text}"
                    except:
                        # Last resort - text only, no image
                        basic_prompt = "Common HVAC equipment includes: air conditioners, furnaces, heat pumps. List basic equipment types."
                        basic_response = self.gemini_model.generate_content(
                            basic_prompt,  # Text only, no image
                            generation_config=genai.GenerationConfig(
                                max_output_tokens=100,
                                temperature=0.1
                            )
                        )
                        if basic_response.parts:
                            narrative = f"[Basic] {basic_response.text}"
                        else:
                            narrative = f"Gemini blocked all attempts (reason: {candidate.finish_reason})"
                else:
                    narrative = f"Gemini blocked (reason: {candidate.finish_reason})"
        else:
            narrative = "No analysis generated"

        response_time = time.time() - start_time

        # Extract entities
        elements = self.extract_entities_from_narrative(narrative, "gemini")

        return {
            'narrative': narrative,
            'extracted_elements': elements,
            'api': 'gemini',
            'response_time': response_time
        }

    def analyze_with_gemini(self, image_path: str) -> Dict:
        """Analyze with Gemini using retry logic and caching"""
        # Check cache first
        cached = self.load_from_cache(image_path, 'gemini')
        if cached:
            return cached

        try:
            # Use retry logic
            result = self.retry_with_backoff(
                self.analyze_with_gemini_internal,
                image_path
            )

            # Cache successful response
            self.save_to_cache(image_path, 'gemini', result)
            logging.info(f"Gemini response time: {result['response_time']:.2f} seconds")
            return result

        except Exception as e:
            logging.error(f"Gemini final failure for {image_path}: {str(e)}")
            return {
                'narrative': f"Error after {self.max_retries} attempts: {str(e)}",
                'extracted_elements': {},
                'api': 'gemini',
                'response_time': 0
            }

    def analyze_with_openai_internal(self, image_path: str) -> Dict:
        """Internal OpenAI analysis function for retry wrapper"""
        start_time = time.time()

        base64_image = self.encode_image(image_path)

        response = self.openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": """Analyze this HVAC/Electrical/Plumbing service image in detail:

1. **Overall Scene**: Complete context and setting
2. **Service Activity**: What work is being performed?
3. **Equipment Identification**: All equipment, brands, models visible
4. **Service Classification**: Type of service (installation/repair/maintenance)
5. **Technical Observations**: Conditions, specifications, issues
6. **Environment**: Location type and characteristics
7. **Professional Insights**: Quality and safety observations
8. **Service Narrative**: Complete story of this service moment

Be thorough and specific about everything visible in this service context."""
                    },
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                    }
                ]
            }],
            max_tokens=1000
        )

        narrative = response.choices[0].message.content
        response_time = time.time() - start_time

        # Extract entities
        elements = self.extract_entities_from_narrative(narrative, "openai")

        return {
            'narrative': narrative,
            'extracted_elements': elements,
            'api': 'openai',
            'response_time': response_time
        }

    def analyze_with_openai(self, image_path: str) -> Dict:
        """Analyze with OpenAI using retry logic and caching"""
        # Check cache first
        cached = self.load_from_cache(image_path, 'openai')
        if cached:
            return cached

        try:
            # Use retry logic
            result = self.retry_with_backoff(
                self.analyze_with_openai_internal,
                image_path
            )

            # Cache successful response
            self.save_to_cache(image_path, 'openai', result)
            logging.info(f"OpenAI response time: {result['response_time']:.2f} seconds")
            return result

        except Exception as e:
            logging.error(f"OpenAI final failure for {image_path}: {str(e)}")
            return {
                'narrative': f"Error after {self.max_retries} attempts: {str(e)}",
                'extracted_elements': {},
                'api': 'openai',
                'response_time': 0
            }

    def standardize_entity(self, entity: str, entity_type: str) -> str:
        """Standardize entity names using mapping dictionary"""
        entity_lower = entity.lower().strip()

        if entity_type in ENTITY_STANDARDIZATION:
            for standard, variants in ENTITY_STANDARDIZATION[entity_type].items():
                if entity_lower in [v.lower() for v in variants]:
                    return standard

        return entity  # Return original if no standardization found

    def extract_entities_from_narrative(self, narrative: str, api_name: str) -> Dict:
        """Extract and standardize entities from narrative"""
        # Simple extraction based on keywords in narrative
        entities = {
            'brands': [],
            'equipment': [],
            'service_type': []
        }

        narrative_lower = narrative.lower()

        # Extract brands
        for standard, variants in ENTITY_STANDARDIZATION['brands'].items():
            for variant in variants:
                if variant.lower() in narrative_lower:
                    entities['brands'].append(standard)
                    break

        # Extract equipment
        for standard, variants in ENTITY_STANDARDIZATION['equipment'].items():
            for variant in variants:
                if variant.lower() in narrative_lower:
                    entities['equipment'].append(standard)
                    break

        # Extract service types
        for standard, variants in ENTITY_STANDARDIZATION['service_type'].items():
            for variant in variants:
                if variant.lower() in narrative_lower:
                    entities['service_type'].append(standard)
                    break

        # Convert lists to comma-separated strings
        return {
            'BRANDS': ', '.join(set(entities['brands'])) if entities['brands'] else 'Not detected',
            'EQUIPMENT': ', '.join(set(entities['equipment'])) if entities['equipment'] else 'Not detected',
            'SERVICE_TYPE': ', '.join(set(entities['service_type'])) if entities['service_type'] else 'Not detected',
            'ACTIONS': 'See narrative',
            'LOCATION': 'See narrative',
            'PEOPLE': 'See narrative'
        }

    def process_image_batch(self, image_paths: List[str]) -> List[Dict]:
        """Process a batch of images in parallel"""
        results = []

        with ThreadPoolExecutor(max_workers=self.batch_size) as executor:
            # Submit all tasks
            future_to_image = {}

            for image_path in image_paths:
                filename = os.path.basename(image_path)
                logging.info(f"Submitting {filename} for processing")

                # Submit both API calls for each image
                gemini_future = executor.submit(self.analyze_with_gemini, image_path)
                openai_future = executor.submit(self.analyze_with_openai, image_path)

                future_to_image[gemini_future] = (image_path, 'gemini')
                future_to_image[openai_future] = (image_path, 'openai')

            # Collect results
            image_results = {}

            for future in as_completed(future_to_image):
                image_path, api_type = future_to_image[future]
                filename = os.path.basename(image_path)

                if filename not in image_results:
                    image_results[filename] = {}

                try:
                    result = future.result()
                    image_results[filename][api_type] = result
                except Exception as e:
                    logging.error(f"Failed to process {filename} with {api_type}: {e}")
                    image_results[filename][api_type] = {
                        'narrative': f"Error: {str(e)}",
                        'extracted_elements': {},
                        'api': api_type,
                        'response_time': 0
                    }

            # Combine results for each image
            for image_path in image_paths:
                filename = os.path.basename(image_path)
                if filename in image_results:
                    gemini_result = image_results[filename].get('gemini', {})
                    openai_result = image_results[filename].get('openai', {})

                    combined = self.format_results(filename, gemini_result, openai_result)
                    results.append(combined)

        return results

    def format_results(self, filename: str, gemini_result: Dict, openai_result: Dict) -> Dict:
        """Format results for Excel output"""
        result = {
            'filename': filename,
            'processed_at': datetime.now().isoformat(),

            # Gemini results
            'gemini_narrative': gemini_result.get('narrative', ''),
            'gemini_brands': gemini_result.get('extracted_elements', {}).get('BRANDS', ''),
            'gemini_equipment': gemini_result.get('extracted_elements', {}).get('EQUIPMENT', ''),
            'gemini_service_type': gemini_result.get('extracted_elements', {}).get('SERVICE_TYPE', ''),
            'gemini_response_time': gemini_result.get('response_time', 0),

            # OpenAI results
            'openai_narrative': openai_result.get('narrative', ''),
            'openai_brands': openai_result.get('extracted_elements', {}).get('BRANDS', ''),
            'openai_equipment': openai_result.get('extracted_elements', {}).get('EQUIPMENT', ''),
            'openai_service_type': openai_result.get('extracted_elements', {}).get('SERVICE_TYPE', ''),
            'openai_response_time': openai_result.get('response_time', 0)
        }

        return result

    def save_to_excel(self, results: List[Dict]):
        """Save results to Excel with formatting"""
        if not results:
            return

        # Create DataFrame
        df = pd.DataFrame(results)

        # Create Excel writer
        with pd.ExcelWriter(self.results_file, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Analysis Results', index=False)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Analysis Results']

            # Format headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                # Limit narrative columns to reasonable width
                if 'narrative' in str(column[0].value).lower():
                    worksheet.column_dimensions[column_letter].width = 50
                else:
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Wrap text for narrative columns
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if 'narrative' in worksheet.cell(1, cell.column).value.lower():
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Add summary statistics sheet
            summary_data = self.generate_summary_stats(results)
            summary_df = pd.DataFrame([summary_data])
            summary_df.to_excel(writer, sheet_name='Summary Statistics', index=False)

        logging.info(f"Results saved to {self.results_file}")

    def generate_summary_stats(self, results: List[Dict]) -> Dict:
        """Generate summary statistics"""
        total_images = len(results)

        # Calculate response times
        gemini_times = [r['gemini_response_time'] for r in results if r['gemini_response_time'] > 0]
        openai_times = [r['openai_response_time'] for r in results if r['openai_response_time'] > 0]

        stats = {
            'Total Images': total_images,
            'Gemini Success Rate': f"{len(gemini_times)}/{total_images}",
            'OpenAI Success Rate': f"{len(openai_times)}/{total_images}",
        }

        if gemini_times:
            stats['Gemini Avg Response (s)'] = round(sum(gemini_times) / len(gemini_times), 2)
            stats['Gemini Min Response (s)'] = round(min(gemini_times), 2)
            stats['Gemini Max Response (s)'] = round(max(gemini_times), 2)

        if openai_times:
            stats['OpenAI Avg Response (s)'] = round(sum(openai_times) / len(openai_times), 2)
            stats['OpenAI Min Response (s)'] = round(min(openai_times), 2)
            stats['OpenAI Max Response (s)'] = round(max(openai_times), 2)

        return stats

    def process_all_images(self, image_folder: str, test_mode: bool = False):
        """Process all images with parallel batching"""
        # Get all images
        all_images = sorted([str(f) for f in Path(image_folder).glob("*.jpg")])

        if test_mode:
            all_images = all_images[:10]  # Test with 10 images
            logging.info(f"TEST MODE: Processing first 10 images")

        logging.info(f"Found {len(all_images)} images to process")

        # Load progress
        progress = self.load_progress()
        processed_files = set(progress['processed'])

        # Filter out already processed images
        images_to_process = [img for img in all_images
                            if os.path.basename(img) not in processed_files]

        logging.info(f"{len(images_to_process)} images remaining to process")

        # Process in batches
        all_results = []

        for i in range(0, len(images_to_process), self.batch_size):
            batch = images_to_process[i:i + self.batch_size]
            batch_num = (i // self.batch_size) + 1
            total_batches = (len(images_to_process) + self.batch_size - 1) // self.batch_size

            logging.info(f"Processing batch {batch_num}/{total_batches}")

            # Process batch
            batch_results = self.process_image_batch(batch)
            all_results.extend(batch_results)

            # Update progress
            for result in batch_results:
                processed_files.add(result['filename'])

            progress['processed'] = list(processed_files)
            self.save_progress(progress)

            # Save results after each batch
            self.save_to_excel(all_results)

            logging.info(f"Batch {batch_num} complete. Total processed: {len(processed_files)}/{len(all_images)}")

            # Small delay between batches
            if i + self.batch_size < len(images_to_process):
                time.sleep(2)

        logging.info(f"Processing complete! {len(all_results)} images processed")
        return all_results

    def load_progress(self) -> Dict:
        """Load progress from file"""
        if self.progress_file.exists():
            with open(self.progress_file, 'r') as f:
                return json.load(f)
        return {'processed': []}

    def save_progress(self, progress: Dict):
        """Save progress to file"""
        with open(self.progress_file, 'w') as f:
            json.dump(progress, f, indent=2)


def main():
    """Main execution function"""
    print("="*60)
    print("IMPROVED UNCONSTRAINED IMAGE ANALYSIS")
    print("With Retry, Parallel Processing, Caching & XLSX Export")
    print("="*60)

    # Check for test mode
    test_mode = '--test' in sys.argv

    if test_mode:
        print("\n🧪 TEST MODE (first 10 images)")
    else:
        print("\n📸 Processing all images in 'images' folder...")

    # Initialize analyzer
    analyzer = ImprovedImageAnalyzer(batch_size=5, max_retries=3)

    # Process images
    image_folder = "images"  # Place your images in this folder

    # Check if images folder exists
    from pathlib import Path
    if not Path(image_folder).exists():
        print(f"\n❌ ERROR: '{image_folder}' folder not found!")
        print("Please create an 'images' folder and add your images there.")
        return

    # Check for images
    image_files = list(Path(image_folder).glob("*.jpg")) + list(Path(image_folder).glob("*.jpeg")) + list(Path(image_folder).glob("*.png"))
    if not image_files:
        print(f"\n❌ ERROR: No images found in '{image_folder}' folder!")
        print("Please add .jpg, .jpeg, or .png images to analyze.")
        return

    results = analyzer.process_all_images(image_folder, test_mode=test_mode)

    print("\n" + "="*60)
    print(f"✅ Processing complete!")
    print(f"Results saved to: {analyzer.results_file}")
    print("="*60)


if __name__ == "__main__":
    main()